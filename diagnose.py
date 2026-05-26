import os
import re
import datetime
from openpyxl import load_workbook

# KONFIGURATION
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_LEAGUES = os.path.join(SCRIPT_DIR, 'Tournament Manager 2023 neu - Ligapyramide 2026.xlsx')
FILE_TEAMS = os.path.join(SCRIPT_DIR, 'Tournament Manager 2023 neu - Vereine Übersicht Ligen Zuordnung.xlsx')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'DIAGNOSE_REPORT.txt')

# HELFER
def clean_id(val):
    if val is None: return None
    s = str(val).strip()
    if "-" in s and any(x in s for x in ["2024", "2025", "2026", "00:00"]):
        try:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    dt = datetime.datetime.strptime(s, fmt)
                    return f"{dt.month}-{dt.day}"
                except: pass
        except: pass
    match = re.match(r'^0*(\d+)-0*(\d+)$', s)
    if match: return f"{match.group(1)}-{match.group(2)}"
    return s

# SPEICHER FÜR FEHLER
errors_duplicates = []
errors_ghosts = []
errors_ids = []

print("--- DIAGNOSE LÄUFT... ---")

# 1. LIGEN LADEN
valid_leagues = {}
try:
    wb = load_workbook(filename=FILE_LEAGUES, data_only=True)
    ws = wb.active 
    start_row = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[4] == "Liga-Nr." or (row[0] and "reine Ligazählung" in str(row[0])):
            start_row = i + 1
            break
            
    for i, row in enumerate(ws.iter_rows(min_row=start_row+1, values_only=True)):
        if len(row) > 5:
            lid = clean_id(row[4])
            if lid: valid_leagues[lid] = row[5]
except Exception as e:
    print(f"❌ Fehler beim Laden der Ligen: {e}")

# 2. VEREINE PRÜFEN
seen_names = {}
try:
    wb = load_workbook(filename=FILE_TEAMS, data_only=True)
    ws = wb.active
    
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        row_num = i + 3 # Excel Zeilennummer
        
        # Leere Zeile Check
        if not any(row): continue 

        raw_name = row[1]
        raw_id = row[7]

        # KATEGORIE: GEISTER-ZEILEN
        if raw_name is None:
            if raw_id is not None:
                errors_ghosts.append(f"Zeile {row_num}: Name leer, aber ID '{raw_id}' vorhanden.")
            continue

        name = str(raw_name).strip()
        if "Koordinaten" in name: continue 

        # KATEGORIE: DUPLIKATE
        if name in seen_names:
            first = seen_names[name]
            errors_duplicates.append(f"Zeile {row_num}: '{name}' (Bereits in Zeile {first})")
        else:
            seen_names[name] = row_num

        # KATEGORIE: ID FEHLER (DATUM)
        if raw_id is not None:
            cleaned = clean_id(raw_id)
            if cleaned not in valid_leagues:
                # Prüfen ob es nach Datum aussieht
                is_date = "-" in str(raw_id) and len(str(raw_id)) > 5
                note = " (Sieht nach Datum aus!)" if is_date else ""
                errors_ids.append(f"Zeile {row_num}: '{name}' hat ID '{raw_id}' -> Unbekannt!{note}")

except Exception as e:
    print(f"❌ Fehler beim Laden der Vereine: {e}")

# 3. REPORT SCHREIBEN
report_lines = []
report_lines.append(f"=== DIAGNOSE BERICHT ({datetime.datetime.now().strftime('%H:%M:%S')}) ===")
report_lines.append(f"Geprüfte Vereine: {len(seen_names)}")
report_lines.append("-" * 50)

if errors_duplicates:
    report_lines.append(f"\n🔴 DOPPELTE VEREINE ({len(errors_duplicates)})")
    report_lines.append("LÖSUNG: Lösche diese Zeilen komplett aus Excel.")
    report_lines.extend([f" - {e}" for e in errors_duplicates])

if errors_ghosts:
    report_lines.append(f"\n👻 GEISTER-ZEILEN ({len(errors_ghosts)})")
    report_lines.append("LÖSUNG: Lösche diese Zeilen (sie sehen leer aus, sind es aber nicht).")
    report_lines.extend([f" - {e}" for e in errors_ghosts])

if errors_ids:
    report_lines.append(f"\n📅 ID / DATUM FEHLER ({len(errors_ids)})")
    report_lines.append("LÖSUNG: Ändere die ID in Excel zu Text (siehe Anleitung).")
    report_lines.extend([f" - {e}" for e in errors_ids])

if not (errors_duplicates or errors_ghosts or errors_ids):
    report_lines.append("\n✅ ALLES SAUBER! Keine Fehler gefunden.")

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write("\n".join(report_lines))

print(f"✅ Diagnose fertig! Öffne Datei: DIAGNOSE_REPORT.txt")